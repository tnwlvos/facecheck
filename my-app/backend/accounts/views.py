from rest_framework.views import APIView # APIView 기반 클래스를 만들기 위한 모듈
from rest_framework.response import Response # RESTful 응답 생성
from rest_framework import status  # HTTP 상태 코드 사용
from django.contrib.auth.models import User # Django 기본 User 모델
from rest_framework_simplejwt.tokens import RefreshToken # JWT 토큰 생성
from rest_framework_simplejwt.views import TokenObtainPairView  # JWT 로그인 뷰
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer # JWT 직렬화기
from rest_framework.decorators import api_view, permission_classes # 함수형 뷰를 위한 데코레이터
from rest_framework.permissions import IsAuthenticated # 인증된 사용자만 접근 허용
from .models import Student # Student 모델
from .models import AttendanceRecord # AttendanceRecord 모델
from io import BytesIO # 파일 작업을 위한 모듈
from django.http import JsonResponse 
from django.http import HttpResponse  # JSON 및 HTTP 응답 생성
from django.utils.timezone import now
from django.utils.timezone import localtime  # 현재 시간 및 지역 시간 함수
from datetime import time # 시간 관련 모듈
import openpyxl # 엑셀 파일 작업 모듈
from face_recognition.real_time_recognition import real_time_face_detection
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.csrf import csrf_exempt
import subprocess
import logging
from django.db.models import Exists, OuterRef, Subquery

# 로깅 설정
logger = logging.getLogger(__name__)

# JWT 토큰에 사용자 커스텀 정보를 추가하기 위한 Serializer
class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['username'] = user.username # JWT에 사용자 이름 추가
        token['name'] = user.first_name  # JWT에 사용자 이름 추가
        return token

    def validate(self, attrs):
        data = super().validate(attrs)
        data['name'] = self.user.first_name # 응답 데이터에 사용자 이름 추가
        return data

# JWT 기반 로그인 뷰 (커스텀 Serializer 사용)
class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

# 회원가입 처리 API 뷰
class RegisterView(APIView):
    def post(self, request):
        # 요청 데이터에서 사용자 정보 가져오기
        username = request.data.get('username')
        password = request.data.get('password')
        name = request.data.get('name')

        # 필수 필드 검증
        if not username or not password or not name:
            return Response({"error": "모든 필드를 입력하세요."}, status=status.HTTP_400_BAD_REQUEST)
        
        #이미 존재하는 사용자 확인
        if User.objects.filter(username=username).exists():
            return Response({"error": "이미 존재하는 사용자입니다."}, status=status.HTTP_400_BAD_REQUEST)
        # 사용자 생성   
        user = User.objects.create_user(username=username, password=password, first_name=name)
        user.save()
        return Response({"message": "회원가입 성공"}, status=status.HTTP_201_CREATED)

# 학생 데이터를 엑셀 파일에서 업로드하는 API
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_students(request):
    user = request.user # 현재 요청한 사용자 가져오기
    if request.method == 'POST' and request.FILES.get('file'): # 파일 업로드 확인
        excel_file = request.FILES['file']
        
        
        try:
            # 엑셀 파일 로드
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # 기존 사용자 학생 데이터 삭제
            Student.objects.filter(user=user).delete()
            success_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):  # 2번째 행부터 데이터 읽기
                try:
                    student_id = str(row[0]).strip()  # 학번
                    name = str(row[1]).strip()  # 이름

                    # 데이터 유효성 검사
                    if not student_id.isdigit() or not name:
                        raise ValueError(f"Invalid data: {row}")

                    # 학생 데이터 생성
                    Student.objects.update_or_create(
                        user=user,
                        student_id=student_id,
                        defaults={'name': name},
                    )
                    success_count += 1

                except ValueError as e:
                    # 잘못된 데이터가 있으면 로그를 출력
                    print(f"⚠️ Invalid row skipped: {row} ({e})")
            
            return Response(
                {"message": f"{success_count}명의 학생 데이터가 업로드되었습니다."},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": f"파일 처리 중 오류가 발생했습니다: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    return Response({"error": "파일이 없습니다."}, status=status.HTTP_400_BAD_REQUEST)



   # 사용자 학생 데이터를 가져오는 API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_students(request):
    user = request.user
    students = Student.objects.filter(user=user).values(
        "student_id", 
        "name",
        "is_present",
        "attendance_time")
    return Response(students, status=status.HTTP_200_OK)


# 특정 학생의 출석을 기록하는 API
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_attendance(request):
    user = request.user
    student_id = request.data.get('student_id')

    if not student_id:
        return JsonResponse({'error': 'Student ID is required'}, status=400)

    try:
        student = Student.objects.get(user=user, student_id=student_id)
        
        #attendance_time = localtime(now())
        AttendanceRecord.objects.create( # 출석 기록 생성
            student=student,
            date=now(),
            time=now().time(),
        )
        # 출석 시간과 상태 업데이트
        student.attendance_time = now()  # 출석 시간 기록
        student.is_present = True       # 출석 상태 True로 설정
        student.save()

        return Response({
            'message': f"{student.name} 출석 처리 완료!",
            'student_id': student.student_id,
            'is_present': student.is_present,
            'attendance_time': student.attendance_time
        }, status=200)
    except Student.DoesNotExist:
        return Response({'error': '학생 정보를 찾을 수 없습니다.'}, status=404)


# 출석 데이터를 엑셀 파일로 내보내는 API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance(request):
    user = request.user  # 요청을 보낸 사용자를 가져옴
    students = Student.objects.filter(user=user)
    # 이전 출석 데이터를 불러옵니다
    attendance_records = AttendanceRecord.objects.filter(student__user=user)
    # 날짜 리스트 생성
    dates = sorted(set(record.date.date() for record in attendance_records))
    print("날짜 리스트:", dates)  # 디버깅용 출력
    
    # 새로운 엑셀 파일 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Records"

    # 헤더 추가
    headers = ["학번","이름"]
    headers.extend([date.strftime('%Y-%m-%d') for date in dates])
    
    
    sheet.append(headers)

     # 학생 데이터 및 출석 시간 추가
    for student in students:
        row = [student.student_id, student.name]
        for date in dates:
            # 해당 날짜의 출석 기록을 확인
            record = attendance_records.filter(student=student, date__date=date).first()
            if record :
                local_time =record.date.strftime('%H:%M')
                print(f"UTC Time: {record.date}, Local Time: {local_time}")
                row.append(local_time)
                
            else:
                row.append("결석")  # 기록이 없으면 결석 처리
        sheet.append(row)
    # 응답 설정
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = f'attachment; filename=attendance_records.xlsx'

    # 엑셀 파일 저장 및 반환
    workbook.save(response)
    return response


@csrf_exempt
def detect_face_api(request):
    try:
        if request.method == 'POST':
            print("🔍 POST 요청 수신!")
            real_time_face_detection()  # 얼굴 인식 실행
            return JsonResponse({"message": "얼굴 인식이 성공적으로 실행되었습니다."})
        else:
            return JsonResponse({"message": "잘못된 요청입니다."}, status=400)
    except Exception as e:
        # 오류 출력
        import traceback
        error_details = traceback.format_exc()
        print(f"🔴 오류 발생: {error_details}")
        return JsonResponse(
            {"message": "얼굴 인식 실행 중 오류가 발생했습니다.", "error": error_details},
            status=500,
        )

@ensure_csrf_cookie
def get_csrf_token(request):
    return JsonResponse({"message": "CSRF cookie set"})

